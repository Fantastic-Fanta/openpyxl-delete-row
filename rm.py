from openpyxl.utils import get_column_letter

def rm_rows(ws, idx, amt):
    rp = {}
    for r in list(ws.row_dimensions.keys()):
        if isinstance(r, int):
            dim = ws.row_dimensions[r]
            rp[r] = {'h': dim.height, 'hid': dim.hidden}
    ws.delete_rows(idx, amt)
    nrp = {}
    for old_r, props in rp.items():
        if old_r < idx:
            nrp[old_r] = props
        elif old_r >= idx + amt:
            new_r = old_r - amt
            nrp[new_r] = props
    for r in range(idx, idx + amt + 100):
        ws.row_dimensions.pop(r, None)
    for r, props in nrp.items():
        dim = ws.row_dimensions[r]
        dim.height = props['h']
        dim.hidden = props['hid']
    new_mc = []
    for m in list(ws.merged_cells.ranges):
        minr = m.min_row
        maxr = m.max_row
        minc = m.min_col
        maxc = m.max_col
        if maxr < idx:
            new_mc.append(m.coord)
            continue
        if minr >= idx and maxr < idx + amt:
            continue
        if minr >= idx + amt:
            shift = -amt
            new_minr = minr + shift
            new_maxr = maxr + shift
            coord = f"{get_column_letter(minc)}{new_minr}:{get_column_letter(maxc)}{new_maxr}"
            new_mc.append(coord)
            continue
        new_maxr = min(maxr, idx - 1)
        if new_maxr >= minr:
            coord = f"{get_column_letter(minc)}{minr}:{get_column_letter(maxc)}{new_maxr}"
            new_mc.append(coord)
    ws.merged_cells.ranges = set()
    for coord in new_mc:
        ws.merge_cells(coord)
    max_r = ws.max_row or 1
    for r in list(ws.row_dimensions.keys()):
        if isinstance(r, int) and r > max_r + 100:
            del ws.row_dimensions[r]
